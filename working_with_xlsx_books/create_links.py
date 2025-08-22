from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# Путь к папке с pdf (рядом со скриптом)
pdf_folder = Path(__file__).parent / "certificates_and_passports"

# Путь к Excel файлу
excel_path = Path(__file__).parent / "Documents_BD.xlsx"

# Загружаем книгу и активный лист
wb = load_workbook(excel_path)
ws = wb.active

# Находим индекс столбца с "document_id"
header_row = 1
doc_id_col = None
for cell in ws[header_row]:
    if cell.value == "document_id":
        doc_id_col = cell.column  # номер столбца (число)
        break

if doc_id_col is None:
    raise ValueError("Столбец 'document_id' не найден в файле Excel")

# Создадим новый столбец справа от document_id для ссылки
link_col = ws.max_column + 1
ws.cell(row=header_row, column=link_col, value="document_link")

# Проходим по строкам с данными
for row in range(header_row + 1, ws.max_row + 1):
    doc_id = ws.cell(row=row, column=doc_id_col).value
    if doc_id:
        pdf_path = pdf_folder / f"{doc_id}.pdf"
        if pdf_path.exists():
            # Добавляем гиперссылку на файл
            link_cell = ws.cell(row=row, column=link_col)
            link_cell.value = "Открыть PDF"
            link_cell.hyperlink = pdf_path.as_uri()
            link_cell.font = Font(color="0000FF", underline="single")
        else:
            # Если файл не найден, можно оставить пусто или написать "Файл не найден"
            ws.cell(row=row, column=link_col, value="Файл не найден")

# Сохраняем изменения
wb.save(excel_path)
print("Гиперссылки добавлены успешно.")
