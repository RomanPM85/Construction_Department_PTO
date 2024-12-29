from pathlib import Path
from openpyxl import load_workbook, Workbook
import os

from openpyxl.styles import PatternFill, Alignment, Font, Border, Side


def merge_xlsx_files(output_filepath, *input_filepaths):
    """Объединяет несколько файлов XLSX в один.

    Args:
        output_filepath: Путь к новому файлу XLSX, в который будут записаны данные.
        *input_filepaths: Список путей к файлам XLSX, которые нужно объединить.
    """

    try:
        workbook = Workbook()
        sheet = workbook.active

        for input_filepath in input_filepaths:
            try:
                input_workbook = load_workbook(input_filepath)
                for sheet_name in input_workbook.sheetnames:
                    input_sheet = input_workbook[sheet_name]
                    # Копируем данные из листа:
                    for row in input_sheet.iter_rows():
                        sheet.append([cell.value for cell in row])
            except FileNotFoundError:
                print(f"Файл {input_filepath} не найден.")
            except Exception as e:
                print(f"Ошибка при обработке файла {input_filepath}: {e}")

        workbook.save(output_filepath)
        print(f"Файлы успешно объединены в {output_filepath}")

    except Exception as e:
        print(f"Произошла общая ошибка: {e}")


def get_filenames_recursive_pathlib(directory):
    """Recursively gets all filenames using pathlib."""
    try:
        path = Path(directory)
        return [x for x in path.rglob('*.xlsx') if x.is_file()]  # Convert Path objects to strings.

    except FileNotFoundError:
        print(f"Error: Directory '{directory}' not found.")
        return []
    except OSError as e:
        print(f"An error occurred: {e}")
        return []


def set_auto_column_width(filepath, sheet_name="Sheet"):
    """Автоматически устанавливает ширину столбцов в файле Excel."""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Буква столбца
            for cell in col:
                try:  # Обработка возможных ошибок при получении значения ячейки
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                except:
                    pass
            adjusted_width = (max_length + 2) * 1.05  # Добавляем 2 символа запаса и коэф-т для ширины
            sheet.column_dimensions[column].width = adjusted_width
        workbook.save(filepath)
    except FileNotFoundError:
        print(f"Файл {filepath} не найден.")
    except Exception as e:
        print(f"Ошибка: {e}")


def add_header_to_excel(filepath, header_row):
    """Добавляет строку заголовков в существующий файл Excel.

    Args:
        filepath: Путь к существующему файлу Excel.
        header_row: Список значений для строки заголовков.
    """
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active
        sheet.insert_rows(1, amount=1)  # Вставляем 1 строку перед первой строкой
        # sheet.append(header_row)
        # sheet[1] = header_row    # Записываем заголовки
        for i, value in enumerate(header_row):
            sheet.cell(row=1, column=i + 1, value=value)
        workbook.save(filepath)
        print(f"Заголовки добавлены к файлу {filepath}.")
    except FileNotFoundError:
        print(f"Файл {filepath} не найден.")
    except Exception as e:
        print(f"Ошибка: {e}")


def format_first_row(filepath):
    """
    Форматирует первую строку листа: закрашивает, центрирует текст, делает жирным.
    """
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active

        """ Получаем количество столбцов в первой строке.  Если лист пустой, 
        это может вызвать ошибку, так что добавим обработку
        """
        try:
            num_cols = len(sheet[1])
        except IndexError:
            print("Лист пустой, форматирование невозможно")
            return

        # Стиль заливки
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Желтый цвет

        # Стиль выравнивания
        alignment = Alignment(horizontal="center", vertical="center")

        # Стиль шрифта
        font = Font(bold=True)

        #  Применение стилей к первой строке
        for cell in sheet[1]:
            cell.fill = fill
            cell.alignment = alignment
            cell.font = font

        workbook.save(filepath)
        print(f"Первая строка файла {filepath} отформатирована.")

    except FileNotFoundError:
        print(f"Файл {filepath} не найден.")
    except Exception as e:
        print(f"Ошибка: {e}")


def add_border_to_data_cells(filepath):
    """
    Добавляет границу ко всем ячейкам с данными в файле Excel.

    Args:
        filepath: Путь к входному файлу Excel (.xlsx).
    """
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active  # Или укажите имя листа, если нужно

        # thin_border = Border(all=Side(style="thin"))
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:  # Проверяем, есть ли данные в ячейке
                    cell.border = thin_border

        workbook.save(filepath)
        print(f"Границы добавлены в файл: {filepath}")

    except FileNotFoundError:
        print(f"Файл не найден: {filepath}")
    except Exception as e:
        print(f"Произошла ошибка: {e}")


if __name__ == "__main__":
    # Пример использования:
    output_file = "merged_file.xlsx"

    directory_path = Path.cwd()
    input_files = get_filenames_recursive_pathlib(directory_path)

    # Проверка существования файлов перед запуском
    if all(os.path.exists(file) for file in input_files):
        merge_xlsx_files(output_file, *input_files)
    else:
        print("Один или несколько файлов не найдены.")

    header = ['date', 'file_name', 'sha256', 'path_file']
    add_header_to_excel(output_file, header)

    set_auto_column_width(output_file)
    add_border_to_data_cells(output_file)
