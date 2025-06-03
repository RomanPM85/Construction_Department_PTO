from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

def get_column_names(ws):
    return [cell.value for cell in ws[1]]

def main():
    files = list(Path('.').glob('*.xlsx'))
    if not files:
        print("В текущей папке нет файлов .xlsx")
        return

    combined_wb = Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Объединение"

    combined_columns = []
    combined_col_index = {}

    current_row = 1

    for file_path in files:
        print(f"Обрабатываю файл {file_path.name}")
        wb = load_workbook(file_path)
        ws = wb.active

        header = get_column_names(ws)

        # Если это первый файл, записываем заголовок и формируем словарь индексов
        if not combined_columns:
            combined_columns = header
            for idx, col_name in enumerate(combined_columns, start=1):
                combined_ws.cell(row=1, column=idx, value=col_name)
            combined_col_index = {name: idx for idx, name in enumerate(combined_columns, start=1)}
            current_row = 2

        # Для последующих файлов проверяем и расширяем общий заголовок, если есть новые столбцы
        else:
            new_cols = [col for col in header if col not in combined_columns]
            if new_cols:
                # Добавляем новые столбцы в combined_columns и ws
                for col in new_cols:
                    combined_columns.append(col)
                # Перезаписываем заголовок в combined_ws
                for idx, col_name in enumerate(combined_columns, start=1):
                    combined_ws.cell(row=1, column=idx, value=col_name)
                combined_col_index = {name: idx for idx, name in enumerate(combined_columns, start=1)}

        # Создаем отображение колонок текущего файла
        header_map = {name: idx for idx, name in enumerate(header)}

        # Копируем данные из текущего файла в combined_ws
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Проверяем, что строка не пустая
            if all(cell.value is None for cell in row):
                continue
            for col_name, dest_col_idx in combined_col_index.items():
                val = None
                if col_name in header_map:
                    val = row[header_map[col_name]].value
                combined_ws.cell(row=current_row, column=dest_col_idx, value=val)
            current_row += 1

    # Применяем перенос текста ко всем ячейкам
    for row in combined_ws.iter_rows(min_row=1, max_row=combined_ws.max_row, max_col=len(combined_columns)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    combined_wb.save("Общий_файл.xlsx")
    print("Объединение завершено. Файл сохранён как 'Общий_файл.xlsx'.")

if __name__ == "__main__":
    main()
