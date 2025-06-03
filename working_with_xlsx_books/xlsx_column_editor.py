from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

def get_column_names(ws):
    return [cell.value for cell in ws[1]]

def apply_column_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def copy_row_values(src_row, dest_ws, dest_row_idx, col_map):
    # Копируем значения из src_row в dest_ws в строку dest_row_idx по col_map
    for col_name, dest_col_idx in col_map.items():
        # Найдем индекс колонки в src_row
        try:
            src_col_idx = src_row['header_map'][col_name]
            val = src_row['row'][src_col_idx].value
        except KeyError:
            val = None
        dest_ws.cell(row=dest_row_idx, column=dest_col_idx, value=val)

def main():
    template_path = Path('template.xlsx')
    if not template_path.exists():
        print('Шаблон template.xlsx не найден')
        return

    template_wb = load_workbook(template_path)
    template_ws = template_wb.active

    template_columns = get_column_names(template_ws)

    # Получаем ширину столбцов из шаблона
    template_widths = {}
    for col_cell in template_ws[1]:
        col_letter = col_cell.column_letter
        template_widths[col_letter] = template_ws.column_dimensions[col_letter].width or 10

    processed_files = []

    # Проходим по всем xlsx файлам в текущей папке, кроме шаблона
    for file_path in Path('.').glob('*.xlsx'):
        if file_path.name == 'template.xlsx':
            continue

        print(f'Обработка файла: {file_path.name}')
        wb = load_workbook(file_path)
        ws = wb.active

        current_columns = get_column_names(ws)

        # Удаляем строки дублирующие заголовок
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            values = [cell.value for cell in row]
            if values == template_columns:
                rows_to_delete.append(row[0].row)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

        # Добавляем недостающие столбцы и переименовываем по шаблону
        col_index_map = {name: idx+1 for idx, name in enumerate(current_columns)}

        for idx, col_name in enumerate(template_columns):
            if col_name not in current_columns:
                ws.cell(row=1, column=ws.max_column+1, value=col_name)
                col_index_map[col_name] = ws.max_column

        # Переименовываем столбцы по шаблону
        for col_idx, col_name in enumerate(template_columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value != col_name:
                cell.value = col_name

        current_columns = get_column_names(ws)
        col_index_map = {name: idx+1 for idx, name in enumerate(current_columns)}

        # Столбец "Раздел проекта"
        if "Раздел проекта" not in col_index_map:
            ws.cell(row=1, column=ws.max_column+1, value="Раздел проекта")
            col_index_map["Раздел проекта"] = ws.max_column

        project_name = file_path.stem.replace('-ВОР', '')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[col_index_map["Раздел проекта"]-1].value = project_name

        # Перенос текста для всех ячеек
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.alignment is None:
                    cell.alignment = Alignment(wrap_text=True)
                else:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=True,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent
                    )

        # Устанавливаем ширину столбцов
        apply_column_widths(ws, template_widths)

        wb.save(file_path)
        print(f'Файл {file_path.name} обработан и сохранен.')
        processed_files.append(file_path)

    # Объединение всех обработанных файлов в один
    print('Объединение файлов в ВОР_Общий.xlsx')
    combined_wb = Workbook()
    combined_ws = combined_wb.active
    combined_ws.title = "Общий"

    # Записываем заголовок из шаблона
    for col_idx, col_name in enumerate(template_columns, start=1):
        combined_ws.cell(row=1, column=col_idx, value=col_name)
    # Добавляем "Раздел проекта" если его нет в шаблоне
    if "Раздел проекта" not in template_columns:
        combined_ws.cell(row=1, column=len(template_columns)+1, value="Раздел проекта")
        combined_columns = template_columns + ["Раздел проекта"]
    else:
        combined_columns = template_columns

    # Создаем словарь для быстрого доступа к индексам столбцов в итоговом файле
    combined_col_index = {name: idx+1 for idx, name in enumerate(combined_columns)}

    current_row = 2
    for file_path in processed_files:
        wb = load_workbook(file_path)
        ws = wb.active
        # Сопоставляем столбцы текущего файла с итоговыми по названию
        header = get_column_names(ws)
        header_map = {name: idx for idx, name in enumerate(header)}

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Пропускаем пустые строки (если есть)
            if all(cell.value is None for cell in row):
                continue
            # Копируем значения по колонкам
            for col_name, dest_col_idx in combined_col_index.items():
                val = None
                if col_name in header_map:
                    val = row[header_map[col_name]].value
                combined_ws.cell(row=current_row, column=dest_col_idx, value=val)
            current_row += 1

    # Применяем перенос текста ко всем ячейкам итогового файла
    for row in combined_ws.iter_rows(min_row=1, max_row=combined_ws.max_row, max_col=combined_ws.max_column):
        for cell in row:
            if cell.alignment is None:
                cell.alignment = Alignment(wrap_text=True)
            else:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )

    # Устанавливаем ширину столбцов как в шаблоне
    apply_column_widths(combined_ws, template_widths)

    combined_wb.save('ВОР_Общий.xlsx')
    print('Файл ВОР_Общий.xlsx создан.')

if __name__ == '__main__':
    main()
