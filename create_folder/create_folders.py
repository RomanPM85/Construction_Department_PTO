import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook import Workbook


FILE_EXCEL = 'list_folders.xlsx'


def delete_folder():
    path = '.'
    dirs = os.listdir(path)
    for item in dirs:
        if os.path.isdir(str(item)):
            shutil.rmtree(os.path.join(item))


def create_folder_from_excel(self):
    wb = load_workbook(FILE_EXCEL)
    ws = wb.active
    for row in ws.values:
        for value in row:
            if os.path.isdir(str(item)):
                Path.mkdir(self / value)


def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "folders"
    wb.save(FILE_EXCEL)


if __name__ == "__main__":
    welcome = (f"Hello world!, это программа для создания папок из списка указанного в файле {FILE_EXCEL}.\n"
               f"GNU GPL (GNU General Public License) Mamchiy Roman https://github.com/RomanPM85")
    outpath = Path.cwd()
    print(welcome)
    user_input = input(f"Список команд: \n"
                       f"Создаст файл {FILE_EXCEL}, для создания списка папок, введите => 1 \n"
                       f"Создаст папки из таблицы файла {FILE_EXCEL}, введите => 2 \n"
                       f"Удалит все папки в {outpath}, введите => 3 \n"
                       f"Введите номер команды => ")
    if user_input == '3':
        verification = input(f"Точно удалить ??, если да, повторите ввод =>")
        if verification == '3':
            delete_folder()
        else:
            pass
    elif user_input == '2':
        create_folder_from_excel(outpath)
    elif user_input == '1':
        create_excel()
    else:
        pass
