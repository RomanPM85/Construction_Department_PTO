import pandas as pd
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# Загрузка JSON-файла
json_file = Path('data.json')  # Укажите имя вашего JSON-файла
with json_file.open('r', encoding='utf-8') as file:
    data = json.load(file)

# Преобразование JSON в DataFrame
# df = pd.DataFrame(data['data'])
df = pd.DataFrame(data)


# Сохранение в Excel с форматированием
output_file = Path('output.xlsx')

# Создание ExcelWriter для сохранения с форматированием
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Форматирование таблицы
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Формат заголовков
    header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical='top')
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Формат данных
    data_alignment = Alignment(wrap_text=True, vertical='top')
    data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Применение форматирования к заголовкам
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.value = value
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # Применение форматирования к данным
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = data_border

    # Авторазмер столбцов
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Получить букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    # Фильтрация и замораживание строк
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = 'A2'

print(f"Файл {output_file} успешно создан.")
