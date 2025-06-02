from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Определяем путь к CSV-файлу
csv_file = Path.cwd() / 'data.csv'

# Читаем данные из CSV-файла
df = pd.read_csv(csv_file)

# Определяем путь для сохранения Excel-файла
excel_file = Path.cwd() / 'output.xlsx'

# Создаем Excel-файл и записываем данные
df.to_excel(excel_file, index=False, sheet_name='Данные', engine='openpyxl')

# Загружаем созданный Excel-файл для дальнейшего форматирования
workbook = load_workbook(excel_file)
worksheet = workbook['Данные']

# Устанавливаем стили для таблицы
# Устанавливаем ширину столбцов (примерные значения, можно изменить)
for i, column_cells in enumerate(worksheet.columns, start=1):
    worksheet.column_dimensions[chr(64 + i)].width = 20

# Устанавливаем стили для заголовков
header_font = Font(bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for cell in worksheet[1]:  # Первая строка - это заголовки
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Устанавливаем границы для всех ячеек
for row in worksheet.iter_rows():
    for cell in row:
        cell.border = thin_border

# Сохраняем изменения в Excel-файле
workbook.save(excel_file)

print(f'Файл успешно сохранен: {excel_file}')
