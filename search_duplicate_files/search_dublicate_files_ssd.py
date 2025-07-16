#!/usr/bin/env python3
import os
import hashlib
from collections import defaultdict
from itertools import islice
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

CHUNK = 1024 * 1024  # 1 MiB – размер блока для чтения
REPORT = "duplicates.xlsx"  # имя итогового Excel-файла

# Список папок для сканирования
SCAN_DIRECTORIES = [
    'folder1/',
    'folder2/',
    'folder3/',
]


def sha256(file_path: Path) -> str:
    """Вычислить SHA-256 хэш файла, читая его порциями."""
    h = hashlib.sha256()
    try:
        with file_path.open("rb") as f:
            for chunk in iter(lambda: f.read(CHUNK), b""):
                h.update(chunk)
        return h.hexdigest()
    except (PermissionError, OSError) as err:
        print(f"Ошибка при чтении файла {file_path}: {err}")
        return ""


def collect_hashes(directories: list[str]) -> dict[str, list[Path]]:
    """Обойти указанные каталоги и сгруппировать файлы по хэшу."""
    hashes: dict[str, list[Path]] = defaultdict(list)

    for dir_name in directories:
        root = Path(dir_name)

        # Проверяем существование директории
        if not root.exists() or not root.is_dir():
            print(f"Директория не найдена или не является папкой: {root}")
            continue

        print(f"Сканирование директории: {root}")

        for dirpath, _, filenames in os.walk(root):
            for name in filenames:
                path = Path(dirpath) / name
                try:
                    # Пропускаем символические ссылки и слишком большие файлы
                    if path.is_symlink() or path.stat().st_size > 1024 * 1024 * 1024:  # > 1 ГБ
                        continue

                    file_hash = sha256(path)
                    if file_hash:
                        hashes[file_hash].append(path)
                except (PermissionError, OSError) as err:
                    print(f"Пропуск {path}: {err}")

    return hashes


def build_rows(duplicates: dict[str, list[Path]]) -> list[list[str]]:
    """Преобразовать словарь дубликатов в строки для отчёта."""
    rows = []
    idx = 1
    for paths in duplicates.values():
        paths_sorted = sorted(str(p) for p in paths)
        rows.append([idx, paths_sorted[0], *paths_sorted[1:]])
        idx += 1
    return rows


def save_to_excel(rows: list[list[str]], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duplicate files"

    # --- Шапка --------------------------------------------------------------
    headers = ["№", "Original", "Duplicate 1", "Duplicate 2", "…"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="C0C0C0")  # серый
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # --- Данные -------------------------------------------------------------
    for row in rows:
        ws.append(row)

    # применяем границы ко всем ячейкам
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # автофильтр
    ws.auto_filter.ref = ws.dimensions
    # авто-подбор ширины колонок
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(8, min(max_len + 2, 80))

    wb.save(filename)
    print(f"Отчёт сохранён в {filename}")


def main():
    print("Начало поиска дубликатов...")

    # Получаем полные пути к директориям относительно текущей рабочей директории
    full_paths = [str(Path.cwd() / dir_name) for dir_name in SCAN_DIRECTORIES]

    hashes = collect_hashes(full_paths)
    duplicates = {h: p for h, p in hashes.items() if len(p) > 1}

    if not duplicates:
        print("Дубликатов не найдено.")
        return

    rows = build_rows(duplicates)
    save_to_excel(rows, REPORT)

    print("Поиск дубликатов завершен.")


if __name__ == "__main__":
    main()
