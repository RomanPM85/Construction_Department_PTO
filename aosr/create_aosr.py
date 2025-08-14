import re
from pathlib import Path
from openpyxl import load_workbook

def load_data_rows(data_path):
    wb = load_workbook(data_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data_rows.append(row_dict)

    return data_rows

def load_documents_data(documents_path):
    wb = load_workbook(documents_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    try:
        idx_id = headers.index("document_id")
        idx_type = headers.index("document_type")
        idx_number = headers.index("document_number")
    except ValueError:
        raise Exception("В файле Documents_BD.xlsx нет нужных столбцов: 'document_id', 'document_type' или 'document_number'")

    documents_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        doc_id = row[idx_id]
        doc_type = row[idx_type]
        doc_number = row[idx_number]
        if doc_id is not None:
            documents_dict[str(doc_id)] = {
                "type": str(doc_type) if doc_type is not None else "",
                "number": str(doc_number) if doc_number is not None else ""
            }

    return documents_dict

def replace_variables_in_template(template_path, data, documents_dict, output_path):
    wb = load_workbook(template_path)
    pattern = re.compile(r'\{\{\s*(\w+)\s*\}\}')

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    matches = pattern.findall(cell.value)
                    if matches:
                        new_value = cell.value
                        for var in matches:
                            if var == "var22" and var in data and data[var]:
                                ids = [x.strip() for x in data[var].split(",") if x.strip()]
                                replaced_values = []
                                for doc_id in ids:
                                    doc_info = documents_dict.get(doc_id, {"type": "", "number": ""})
                                    replaced_values.append(f"{doc_info['type']}{doc_info['number']}")
                                replacement_str = ", ".join(replaced_values)
                                new_value = new_value.replace(f"{{{{ {var} }}}}", replacement_str)
                                new_value = new_value.replace(f"{{{{{var}}}}}", replacement_str)
                            elif var in data and data[var] is not None:
                                new_value = new_value.replace(f"{{{{ {var} }}}}", str(data[var]))
                                new_value = new_value.replace(f"{{{{{var}}}}}", str(data[var]))
                        cell.value = new_value

    wb.save(output_path)
    print(f"Файл сохранён: {output_path}")

if __name__ == "__main__":
    data_file = Path("variable_data.xlsx")
    documents_file = Path("Documents_BD.xlsx")
    template_file = Path("template.xlsx")
    output_folder = Path("output_files")  # Папка для сохранения результатов
    output_template = "filled_document_{}.xlsx"

    # Создаём папку, если её нет
    output_folder.mkdir(parents=True, exist_ok=True)

    data_rows = load_data_rows(data_file)
    documents_dict = load_documents_data(documents_file)

    for idx, data_dict in enumerate(data_rows, start=1):
        output_file = output_folder / output_template.format(idx)
        replace_variables_in_template(template_file, data_dict, documents_dict, output_file)
