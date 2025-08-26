from pathlib import Path
from openpyxl import load_workbook


def read_excel_data(file_path, sheet_name=None):
    """Чтение данных из Excel файла."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = []
    headers = {cell.value: col for col, cell in enumerate(ws[1], start=1)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {header: row[col - 1] for header, col in headers.items()}
        data.append(row_data)
    return data, headers


def write_to_variable_data(variable_data_path, updated_rows):
    """Запись обновленных данных в файл variable_data.xlsx."""
    wb = load_workbook(variable_data_path)
    ws = wb.active
    headers = {cell.value: col for col, cell in enumerate(ws[1], start=1)}

    for row_index, row_data in enumerate(updated_rows, start=2):
        for col_name, value in row_data.items():
            if col_name in headers:
                ws.cell(row=row_index, column=headers[col_name]).value = value

    wb.save(variable_data_path)
    print(f"Обновленные данные записаны в файл: {variable_data_path}")


def generate_variable_data():
    # Пути к файлам
    variable_data_path = Path("variable_data.xlsx")
    composition_id_path = Path("Technical_data.xlsx")
    certificates_path = Path("certificates_and_passports.xlsx")

    # Чтение данных из файлов
    variable_data, variable_headers = read_excel_data(variable_data_path)
    composition_id_data, composition_headers = read_excel_data(composition_id_path)
    certificates_data, certificates_headers = read_excel_data(certificates_path)

    # Индексация данных для быстрого доступа
    composition_id_index = {row["document_id"]: row for row in composition_id_data}
    certificates_index = {row["document_id"]: row for row in certificates_data}

    # Обновление данных в variable_data.xlsx
    updated_rows = []
    for row in variable_data:
        document_id = row.get("document_id")
        if not document_id:
            continue

        # Получение данных из Состав_ИД_ФОК.xlsx
        composition_row = composition_id_index.get(document_id, {})
        var8 = composition_row.get("AOSR_number", "")
        var19 = ", ".join(
            filter(None, [
                composition_row.get("Наименование СМР"),
                composition_row.get("Тип"),
                composition_row.get("этаж"),
                composition_row.get("rooms"),
                composition_row.get("оси"),
            ])
        )
        var20 = composition_row.get("шифр проекта", "")
        var21 = composition_row.get("полное название проекта", "")
        var23 = f"{composition_row.get('шифр проекта', '')} {composition_row.get('лист РД', '')}"
        var30 = ", ".join(filter(None, [
            composition_row.get("шифр проекта", ""),
            composition_row.get("Нормативные ссылки", "")
        ]))

        # Корректировка для var31
        var31_raw = composition_row.get("Последующие работы", "")
        var31_parts = [part.strip() for part in var31_raw.split(",")]  # Разделение данных по запятой
        var31_result = []

        for part in var31_parts:
            # Проверка, является ли часть ключом
            var31_data = composition_id_index.get(part, {}).get("Наименование СМР", "")
            if var31_data:
                var31_result.append(var31_data)  # Если ключ найден, добавляем значение
            else:
                var31_result.append(part)  # Если ключ не найден, добавляем текст напрямую

        var31 = ", ".join(var31_result)  # Объединение обработанных частей

        var34 = f"{composition_row.get('шифр проекта', '')} {composition_row.get('лист РД', '')}"

        # Получение данных для столбца var22
        var22_materials = composition_row.get("Наименование материалов", "")
        var22_certificates = composition_row.get("sertificat_id", "")
        var22_certificate_values = []  # Список значений document_type + document_number
        if var22_certificates:
            for cert_id in map(str.strip, var22_certificates.split(",")):
                certificate_row = certificates_index.get(cert_id, {})
                document_type = certificate_row.get("document_type", "")
                document_number = certificate_row.get("document_number", "")
                if document_type or document_number:
                    var22_certificate_values.append(f"{document_type} {document_number}")
        var22 = ", ".join(filter(None, [var22_materials, *var22_certificate_values]))

        # Формирование строки для обновления
        updated_row = {
            "var8": var8,
            "var19": var19,
            "var20": var20,
            "var21": var21,
            "var22": var22,
            "var23": var23,
            "var30": var30,
            "var31": var31,
            "var34": var34,
        }

        updated_rows.append(updated_row)

    # Запись обновленных данных в variable_data.xlsx
    write_to_variable_data(variable_data_path, updated_rows)


if __name__ == "__main__":
    generate_variable_data()
