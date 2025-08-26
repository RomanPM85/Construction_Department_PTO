from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
from pathlib import Path

def get_headers(ws):
    """Получение заголовков из первой строки таблицы."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[header] = col
    return headers

def load_variable_rows(data_filename):
    """Загрузка данных из файла variable_data.xlsx."""
    wb = load_workbook(data_filename, data_only=True)
    ws = wb.active
    headers = get_headers(ws)

    # Проверка на наличие обязательных столбцов
    if "name_files" not in headers or "document_id" not in headers:
        raise ValueError("В файле variable_data.xlsx отсутствуют обязательные столбцы 'name_files' или 'document_id'")

    data_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = {}
        for var_name, col in headers.items():
            cell_value = ws.cell(row=row, column=col).value
            row_data[var_name] = cell_value

        # Формирование имени выходного файла
        name_file = row_data.get("name_files")
        document_id = row_data.get("document_id")

        if name_file and str(name_file).strip():
            filename = str(name_file).strip()
            if not filename.lower().endswith(".xlsx"):
                filename += ".xlsx"
        elif document_id and str(document_id).strip():
            filename = f"{str(document_id).strip()}.xlsx"
        else:
            filename = f"output_row{row-1}.xlsx"

        row_data["_filename"] = filename
        data_rows.append(row_data)

    return data_rows

def replace_variables_in_sheet(ws, variables):
    """Замена переменных в ячейках листа."""
    pattern = re.compile(r"\{\{\s*(\w+)\s*\}\}")

    for row in ws.iter_rows():
        for cell in row:
            # Пропускаем ячейки без текста
            if not cell.value or not isinstance(cell.value, str):
                continue

            # Поиск переменных в ячейке
            matches = pattern.findall(cell.value)
            if not matches:
                continue  # Если переменных нет, пропускаем

            new_value = cell.value
            replaced_any = False  # Флаг, показывающий, была ли замена

            for var in matches:
                val = variables.get(var)
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    replacement = ""
                else:
                    replacement = str(val)
                if f"{{{{ {var} }}}}" in new_value or f"{{{{{var}}}}}" in new_value:
                    replaced_any = True
                new_value = new_value.replace(f"{{{{ {var} }}}}", replacement)
                new_value = new_value.replace(f"{{{{{var}}}}}", replacement)

            # Если была замена, обновляем значение и выравнивание
            if replaced_any:
                cell.value = new_value

                old_alignment = cell.alignment
                cell.alignment = Alignment(
                    horizontal=old_alignment.horizontal,
                    vertical=old_alignment.vertical,
                    text_rotation=old_alignment.textRotation,
                    wrap_text=True,  # Включаем перенос текста только для ячеек с заменой
                    shrink_to_fit=old_alignment.shrinkToFit,
                    indent=old_alignment.indent
                )

def disable_wrap_text(ws, start_col, end_col):
    """Отключение переноса текста для столбцов от A до AN."""
    for col in range(start_col, end_col + 1):
        for row in ws.iter_rows():
            cell = row[col - 1]  # Индексация с 0
            old_alignment = cell.alignment
            cell.alignment = Alignment(
                horizontal=old_alignment.horizontal,
                vertical=old_alignment.vertical,
                text_rotation=old_alignment.textRotation,
                wrap_text=False,  # Отключаем перенос текста
                shrink_to_fit=old_alignment.shrinkToFit,
                indent=old_alignment.indent
            )

def main():
    template_file = Path("template.xlsx")
    data_file = Path("variable_data.xlsx")
    output_dir = Path("output_files_xlsx")

    # Создаём директорию для выходных файлов, если её нет
    output_dir.mkdir(exist_ok=True)

    # Загружаем данные из файла variable_data.xlsx
    data_rows = load_variable_rows(data_file)

    for row_data in data_rows:
        # Загружаем шаблон
        wb = load_workbook(template_file)
        ws = wb.active

        # Отключаем перенос текста для столбцов A-AN
        # disable_wrap_text(ws, start_col=1, end_col=40)  # A-AN (1-40)

        # Заменяем переменные
        replace_variables_in_sheet(ws, row_data)

        # Сохраняем файл
        output_path = output_dir / row_data["_filename"]
        wb.save(output_path)
        print(f"Сохранён файл: {output_path}")

if __name__ == "__main__":
    main()
