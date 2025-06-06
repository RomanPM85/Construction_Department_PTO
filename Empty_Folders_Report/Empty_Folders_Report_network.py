from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_empty_folders_recursively(root: Path):
    empty_folders = []
    for folder in root.rglob('*'):
        if folder.is_dir():
            # Проверяем, что папка пуста (нет файлов и папок)
            if not any(folder.iterdir()):
                empty_folders.append(folder)
    return empty_folders

def safe_resolve(path: Path) -> str:
    try:
        return str(path.resolve())
    except OSError:
        return str(path.absolute())

def create_report(folders, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Пустые папки"

    headers = ["№", "Название папки", "Путь к папке"]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Границы
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Применяем стили к заголовкам
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Заполнение данных и добавление границ
    for i, folder_path in enumerate(folders, 1):
        row_num = i + 1
        abs_path = safe_resolve(folder_path)
        values = [i, folder_path.name, abs_path]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

    # Автоширина столбцов
    for col in range(1, 4):
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Добавление автофильтра по заголовкам
    ws.auto_filter.ref = f"A1:C{len(folders)+1}"

    wb.save(output_file)
    print(f"Отчёт сохранён в {output_file}")

if __name__ == "__main__":
    root_path = Path(".")
    empty_folders = get_empty_folders_recursively(root_path)
    create_report(empty_folders, "empty_folders_report.xlsx")
