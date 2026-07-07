import datetime
import hashlib
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# --- Блок 1: Работа с файловой системой ---

def get_filenames_recursive(directory):
    """Рекурсивно получает все файлы во всех вложенных папках."""
    try:
        path = Path(directory)
        return [x for x in path.rglob('*') if x.is_file()]
    except FileNotFoundError:
        print(f"Ошибка: Директория '{directory}' не найдена.")
        return []
    except OSError as e:
        print(f"Произошла ошибка при чтении папок: {e}")
        return []


def get_file_modification_date(file_path: Path):
    """Возвращает дату модификации файла в читаемом формате."""
    try:
        modification_time = file_path.stat().st_mtime
        return datetime.datetime.fromtimestamp(modification_time).strftime('%Y-%m-%d %H:%M:%S')
    except OSError:
        return "Неизвестно"


def returns_hash_file(file_path: Path, base_directory: Path):
    """Возвращает метаданные файла, его SHA-256 хеш и чистый относительный путь Windows."""
    write_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    modification_date = get_file_modification_date(file_path)

    sha256_hash = hashlib.sha256()
    try:
        with open(file_path, 'rb') as f:
            while chunk := f.read(65536):
                sha256_hash.update(chunk)
        hash_str = sha256_hash.hexdigest()
    except OSError:
        hash_str = "Ошибка чтения файла"

    try:
        rel_path = str(file_path.relative_to(base_directory))
    except ValueError:
        rel_path = file_path.name

    # Принудительно используем классический обратный слэш Windows для красоты
    win_relative_path = rel_path.replace("/", "\\")
    file_extension = file_path.suffix.lower()  # Получаем расширение (например, .txt)

    return write_date, modification_date, file_path.name, file_extension, hash_str, win_relative_path


# --- Блок 2: Форматирование таблицы Excel ---

def process_and_format_excel(filepath, header_row):
    """Применяет профессиональное оформление. Все значения остаются обычным текстом."""
    try:
        workbook = load_workbook(filepath)
        sheet = workbook.active

        # 1. Вставляем строку заголовков в самый верх
        sheet.insert_rows(1, amount=1)
        for col_idx, value in enumerate(header_row, start=1):
            sheet.cell(row=1, column=col_idx, value=value)

        # Включаем автофильтр по границам таблицы
        sheet.auto_filter.ref = sheet.dimensions

        # 2. Подготовка палитры оформления (Стильный синий корпоративный дизайн)
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Темно-синий
        fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Светло-серый

        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")  # Белый жирный текст
        font_data = Font(name="Calibri", size=11, bold=False, color="000000")

        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_data_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

        grid_side = Side(style='thin', color='D9D9D9')
        border_cell = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)

        # Закрепляем первую строку, чтобы она оставалась на экране при прокрутке
        sheet.freeze_panes = 'A2'
        sheet.row_dimensions[1].height = 28

        # 3. Настраиваем и красим только Шапку таблицы (Строка 1)
        for col_idx in range(1, len(header_row) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_header
            cell.border = border_cell

        # 4. Проходим построчно и накладываем стили на ячейки данных (Строка 2+)
        for row_idx in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = 20
            is_even = (row_idx % 2 == 0)

            for col_idx in range(1, len(header_row) + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = border_cell
                cell.font = font_data

                # Чередование цвета строк ("зебра")
                if is_even:
                    cell.fill = fill_zebra

                # Выравнивание контента в зависимости от столбца
                # 1 (Дата), 2 (Дата изм.), 4 (Расширение) центрируем. Хеш (5) и Путь (6) - влево.
                if col_idx in (1, 2, 4):
                    cell.alignment = align_data_center
                else:
                    cell.alignment = align_data_left

                # Явно выставляем текстовый формат ячеек для дат, чтобы Excel не искажал их
                if col_idx in (1, 2):
                    cell.number_format = '@'

        # 5. Авторасширение столбцов по фактической ширине текста
        for col_idx in range(1, sheet.max_column + 1):
            max_length = max(
                (len(str(sheet.cell(row=r, column=col_idx).value or '')) for r in range(1, sheet.max_row + 1)),
                default=0
            )
            column_letter = get_column_letter(col_idx)
            # Добавляем 5 символов запаса (особенно важно для кнопок автофильтра в шапке)
            adjusted_width = max(max_length + 5, 14)
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filepath)
        print(f"Файл {filepath.name} успешно отформатирован и сохранен.")
    except Exception as e:
        print(f"Произошла ошибка при форматировании Excel: {e}")


# --- Главный модуль управления ---

if __name__ == "__main__":
    welcome = (
        "Hi, my name is Roman, this program is designed to get a hash of\n"
        "files written to an xlsx file \n"
        "(The GNU General Public License v3.0) Mamchiy Roman\n"
        "https://github.com\n"
    )
    print(welcome)

    # Выбор режима ввода директорий
    print("Выберите режим работы:")
    print("1 - Сканировать только текущую папку")
    print("2 - Указать пути к нескольким папкам вручную (через запятую)")
    print("3 - Загрузить список путей из внешнего файла 'paths.txt'")

    choice = input("Введите цифру 1, 2 или 3: ").strip()
    directories_to_scan = []

    if choice == "1":
        directories_to_scan.append(Path.cwd())
        print(f"Выбрана текущая рабочая директория: {Path.cwd()}")

    elif choice == "2":
        user_input = input("Введите полные пути к целевым папкам через запятую:\n").strip()
        raw_paths = [p.strip() for p in user_input.split(",") if p.strip()]
        for p in raw_paths:
            path_obj = Path(p)
            if path_obj.is_dir():
                directories_to_scan.append(path_obj)
            else:
                print(f"Предупреждение: Путь '{p}' не существует или не является папкой. Пропущен.")

    elif choice == "3":
        txt_file = Path("paths.txt")
        if not txt_file.exists():
            with open(txt_file, "w", encoding="utf-8") as f:
                f.write("# Вставьте сюда пути к папкам, каждый путь с новой строки\n")
            print("Ошибка: Файл 'paths.txt' не найден. Создан пустой файл шаблона.")
            print("Пожалуйста, заполните его путями и запустите скрипт снова.")
            exit()

        with open(txt_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    path_obj = Path(line)
                    if path_obj.is_dir():
                        directories_to_scan.append(path_obj)
                    else:
                        print(f"Предупреждение: Путь '{line}' не существует или не является папкой. Пропущен.")
    else:
        print("Ошибка: Неверный выбор режима.")
        exit()

    if not directories_to_scan:
        print("Ошибка: Нет доступных директорий для обработки.")
        exit()

    # Общая структура заголовков для всех отчетов
    header_ru = [
        'Дата обновления',
        'Дата изменения файла',
        'Имя файла',
        'Расширение файла',
        'Хеш SHA-256',
        'Относительный путь к файлу'
    ]

    print("\nНачало раздельной обработки папок...")

    # Основной цикл по всем выбранным папкам
    for directory_path in directories_to_scan:
        print(f"\n[Папка] Обработка директории: {directory_path}")

        # Получаем список файлов для текущей папки
        all_files = get_filenames_recursive(directory_path)

        # Формируем имя реестра для текущей папки
        output_filename = f'Реестр_папки_{directory_path.name}.xlsx'
        output_filepath = directory_path / output_filename

        # Исключаем файл локального реестра из обработки, чтобы избежать зацикливания
        all_files = [f for f in all_files if f != output_filepath]

        if not all_files:
            print(f"В папке '{directory_path.name}' нет файлов для обработки.")
            continue

        wb = Workbook()
        ws = wb.active

        # Перезаписываем старый файл реестра, если он уже существовал в этой папке
        output_filepath.unlink(missing_ok=True)

        print(f"Сканирование файлов (всего объектов: {len(all_files)})...")
        for file in all_files:
            file_info = returns_hash_file(file, directory_path)
            ws.append(list(file_info))
            print(f"Успешно обработан: {file.name}")

        wb.save(output_filepath)

        # Запуск форматирования созданного файла
        process_and_format_excel(output_filepath, header_ru)
        print(f"Готово! Реестр сохранен в: {output_filepath}")
    print("\nВсе указанные папки успешно обработаны!")

