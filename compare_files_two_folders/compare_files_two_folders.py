import hashlib
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

def sha256sum(filepath):
    """
    Вычисляет SHA256 хеш файла по заданному пути.
    :param filepath: путь к файлу
    :return: строка с SHA256 хешем
    """
    h = hashlib.sha256()
    # Открываем файл в бинарном режиме и читаем по частям (по 8192 байт),
    # чтобы не загружать весь файл в память сразу
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def collect_files_info(folder):
    """
    Собирает информацию о всех файлах в папке и её подпапках.
    :param folder: объект Path к папке
    :return: список словарей с информацией о файлах:
             путь к папке, имя файла, SHA256, полный путь
    """
    files_info = []
    # Рекурсивно проходим по всем файлам в папке
    for file_path in folder.rglob('*'):
        if file_path.is_file():
            files_info.append({
                "path": str(file_path.parent),  # папка файла
                "name": file_path.name,          # имя файла
                "sha256": sha256sum(file_path), # SHA256 хеш
                "full_path": str(file_path)      # полный путь к файлу
            })
    return files_info

def compare_files_by_hash(files1, files2):
    """
    Сравнивает два списка файлов по их SHA256 хешам.
    Формирует объединённый список с данными из обеих папок.
    :param files1: список файлов из первой папки
    :param files2: список файлов из второй папки
    :return: список списков - строк для итоговой таблицы
    """
    # Создаём словари: хеш -> информация о файле
    hash_to_file1 = {f["sha256"]: f for f in files1}
    hash_to_file2 = {f["sha256"]: f for f in files2}

    # Объединяем все уникальные хеши из обеих папок
    all_hashes = set(hash_to_file1.keys()) | set(hash_to_file2.keys())

    rows = []
    # Для каждого хеша формируем строку с данными из обеих папок
    for h in all_hashes:
        f1 = hash_to_file1.get(h)
        f2 = hash_to_file2.get(h)
        row = [
            f1["path"] if f1 else "",   # путь из первой папки или пусто
            f1["name"] if f1 else "",   # имя файла из первой папки или пусто
            f1["sha256"] if f1 else "", # хеш из первой папки или пусто
            f2["path"] if f2 else "",   # путь из второй папки или пусто
            f2["name"] if f2 else "",   # имя файла из второй папки или пусто
            f2["sha256"] if f2 else "", # хеш из второй папки или пусто
        ]
        rows.append(row)
    return rows

def autofit_columns(ws):
    """
    Автоматически подгоняет ширину колонок листа под содержимое.
    :param ws: лист openpyxl
    """
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # буква колонки, например 'A'
        for cell in col:
            try:
                if cell.value:
                    # Определяем максимальную длину текста в колонке
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Устанавливаем ширину с небольшим запасом
        ws.column_dimensions[column].width = max_length + 2

def style_worksheet(ws):
    """
    Применяет стили к листу: заголовки, границы и автофильтр.
    :param ws: лист openpyxl
    """
    # Шрифт и заливка для заголовков: белый жирный текст на сером фоне
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="808080")

    # Тонкие рамки для ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем стиль к заголовкам (первой строке)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Применяем рамки ко всем ячейкам листа
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Включаем автофильтр для всей области с данными
    ws.auto_filter.ref = ws.dimensions

if __name__ == "__main__":
    # Укажите пути к сравниваемым папкам
    folder1 = Path(r"C:\path\to\first_folder")
    folder2 = Path(r"C:\path\to\second_folder")

    # Собираем информацию о файлах из обеих папок
    files1 = collect_files_info(folder1)
    files2 = collect_files_info(folder2)

    # Сравниваем файлы по хешам и формируем данные для таблицы
    rows = compare_files_by_hash(files1, files2)

    # Создаём новую книгу Excel и лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # Заголовки столбцов
    headers = [
        "Path папки 1", "Имя файла 1", "SHA256 файла 1",
        "Path папки 2", "Имя файла 2", "SHA256 файла 2"
    ]
    ws.append(headers)

    # Добавляем строки с данными
    for row in rows:
        ws.append(row)

    # Подгоняем ширину колонок и применяем стили
    autofit_columns(ws)
    style_worksheet(ws)

    # Сохраняем файл Excel
    wb.save("comparison.xlsx")
    print("Готово! Результат в файле comparison.xlsx")
