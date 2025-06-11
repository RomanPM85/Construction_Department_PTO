import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os


def process_duplicates(file_path):
    # Загружаем файл Excel
    xlsx_path = Path(file_path)
    if not xlsx_path.is_file():
        print(f"Файл {xlsx_path} не найден.")
        return

    try:
        # Проверка, занят ли файл
        with open(xlsx_path, 'r+') as file:
            pass
    except PermissionError:
        print(f"Ошибка: Файл {xlsx_path} занят. Закройте файл и повторите попытку.")
        return

    try:
        # Открываем рабочую книгу и выбираем листы
        wb = load_workbook(xlsx_path)
        ws_duplicates = wb['Duplicate_files']  # Рабочий лист 'Duplicate_files'
        ws_delete = wb['Delete']  # Рабочий лист 'Delete'

        # Читаем данные в DataFrame
        df = pd.read_excel(xlsx_path, sheet_name='Duplicate_files', header=0)

        # Получаем все столбцы, начиная с "B" (то есть "Duplicate 1", "Duplicate 2" и т.д.)
        duplicate_columns = [col for col in df.columns if 'Duplicate' in col]

        # Цвет для заливки - красный
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Список для хранения путей, которые нужно скопировать в лист Delete
        paths_to_delete = []

        # Перебираем строки
        for row_idx, row in df.iterrows():
            # Получаем все пути в текущей строке (не NaN значения)
            paths = [row[col] for col in duplicate_columns if pd.notna(row[col])]

            # Пропускаем, если дубликат только один
            if len(paths) <= 1:
                continue

            # Проверка наличия путей с "%TEMP\"
            temp_paths = []
            non_temp_paths = []

            for i, path_str in enumerate(paths):
                path = Path(str(path_str))  # Преобразуем путь в объект Path
                if "%TEMP\\" in str(path):
                    temp_paths.append(i)
                else:
                    non_temp_paths.append(i)

            # Логика обработки:
            # 1. Если есть и TEMP, и не-TEMP пути - выделяем все TEMP пути
            # 2. Если все пути содержат %TEMP% - оставляем последний
            if non_temp_paths:
                # Есть пути без %TEMP% - выделяем все TEMP пути
                for i in temp_paths:
                    col_name = duplicate_columns[i]
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    cell = ws_duplicates[f"{col_letter}{row_idx + 2}"]
                    cell.fill = red_fill

                    # Добавляем путь в список для копирования
                    paths_to_delete.append(paths[i])

            elif len(temp_paths) > 1:
                # Все пути содержат %TEMP%, оставляем последний
                for i in temp_paths[:-1]:
                    col_name = duplicate_columns[i]
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    cell = ws_duplicates[f"{col_letter}{row_idx + 2}"]
                    cell.fill = red_fill

                    # Добавляем путь в список для копирования
                    paths_to_delete.append(paths[i])

        # Очищаем лист "Delete" перед копированием данных
        for row in range(ws_delete.max_row, 1, -1):
            ws_delete.delete_rows(row)

        # Копируем пути в лист "Delete" в столбец A
        for idx, path in enumerate(paths_to_delete, start=2):  # Начинаем с 2-й строки
            ws_delete[f"A{idx}"] = path

        # Сохраняем изменения в файл
        wb.save(xlsx_path)
        print(f"Обработка завершена. Файл сохранен: {xlsx_path}")
        print(f"Скопировано {len(paths_to_delete)} путей в лист 'Delete'.")

    except PermissionError:
        print(f"Ошибка: Не удается сохранить файл {xlsx_path}. Убедитесь, что файл не открыт в другой программе.")
    except Exception as e:
        print(f"Произошла ошибка: {e}")


if __name__ == "__main__":
    process_duplicates("duplicates_files_report.xlsx")
