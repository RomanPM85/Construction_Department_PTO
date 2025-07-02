import shutil
from pathlib import Path
import openpyxl
import os


def copy_files_to_destination(excel_path, destination_folder):
    """
    Копирует файлы, указанные в столбце Path Excel-файла, в указанную папку назначения.

    :param excel_path: Путь к Excel-файлу
    :param destination_folder: Путь к папке назначения для копирования
    """
    # Преобразуем пути в объекты Path
    destination_path = Path(destination_folder)

    # Создаем папку назначения, если она не существует
    destination_path.mkdir(parents=True, exist_ok=True)

    # Открываем Excel-файл
    workbook = openpyxl.load_workbook(excel_path, read_only=False)

    # Выбираем лист Check
    sheet = workbook['Check']

    # Счетчики для статистики
    total_files = 0
    copied_files = 0
    skipped_files = 0

    # Проходим по строкам, начиная со второй (предполагаем, что первая - заголовки)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        path_cell = row[0]

        # Получаем путь из столбца A
        file_path = path_cell.value

        # Пропускаем пустые строки
        if not file_path:
            continue

        total_files += 1

        # Преобразуем путь с помощью Pathlib
        source_path = Path(file_path)

        # Проверяем существование исходного файла
        if not source_path.exists():
            print(f"Файл не найден: {source_path}")
            skipped_files += 1
            continue

        try:
            # Формируем путь назначения - сохраняем оригинальное имя файла
            destination_file_path = destination_path / source_path.name

            # Копируем файл
            shutil.copy2(source_path, destination_file_path)
            copied_files += 1
            print(f"Скопирован файл: {source_path} -> {destination_file_path}")

        except PermissionError:
            print(f"Ошибка доступа при копировании: {source_path}")
            skipped_files += 1
        except Exception as e:
            print(f"Ошибка при копировании {source_path}: {e}")
            skipped_files += 1

    # Выводим статистику
    print("\nСтатистика копирования:")
    print(f"Всего файлов в списке: {total_files}")
    print(f"Успешно скопировано: {copied_files}")
    print(f"Пропущено файлов: {skipped_files}")


# Пути к файлам
excel_file_path = 'duplicates_files_report_new.xlsx'
destination_folder = r'O:\tcsm_pro\Check'

# Запуск копирования
copy_files_to_destination(excel_file_path, destination_folder)
